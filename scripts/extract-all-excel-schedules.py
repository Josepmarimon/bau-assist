#!/usr/bin/env python3
"""
Extract comprehensive schedule data from all Excel files in the excelhorari directory.
This script extracts: Grado, Curso, Itinerario, Asignatura, Grupo, Aulas, Profesor
and excludes placeholder subjects (white background cells).
"""

import pandas as pd
import numpy as np
import os
import re
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from collections import defaultdict

# Directory with Excel files
EXCEL_DIR = '/Users/josepmarimon/Documents/github/bau-assist/csv/excelhorari'
OUTPUT_FILE = '/Users/josepmarimon/Documents/github/bau-assist/csv/extracted_excel_data.json'

def is_white_background(cell):
    """Check if a cell has white background (placeholder)"""
    if cell.fill and cell.fill.patternType:
        if cell.fill.patternType == 'solid':
            if cell.fill.fgColor:
                color = cell.fill.fgColor.rgb
                if color and (color == 'FFFFFFFF' or color == '00FFFFFF' or color == 'FFFFFF'):
                    return True
    return False

def parse_filename(filename):
    """Extract metadata from filename"""
    info = {}
    
    # Determine degree
    if 'GDIS' in filename:
        info['grado'] = 'Disseny'
        info['grado_code'] = 'GDIS'
    elif 'GBA' in filename:
        info['grado'] = 'Belles Arts'
        info['grado_code'] = 'GBA'
    
    # Extract year
    if '_1r_' in filename:
        info['curso'] = '1r'
    elif '_2n_' in filename:
        info['curso'] = '2n'
    elif '_3r_' in filename:
        info['curso'] = '3r'
    elif '_4t_' in filename:
        info['curso'] = '4t'
    elif '3r_4t_Optativitat' in filename:
        info['curso'] = '3r-4t'
        info['tipo'] = 'Optatives'
    
    return info

def extract_itinerari_from_context(ws, row, col, subject_name):
    """Try to extract itinerari from context around the subject"""
    itinerari = None
    
    # Check for specific subject patterns that indicate itinerari
    itinerari_patterns = {
        'Videojocs': ['Videojocs', 'Game', 'Jocs'],
        'Animació': ['Animació', 'Animation', 'Animació Digital'],
        'Audiovisual': ['Audiovisual', 'Vídeo', 'Cinema'],
        'Gràfic': ['Gràfic', 'Comunicació Visual', 'Disseny Gràfic'],
        'Moda': ['Moda', 'Fashion'],
        'Interiors': ['Interiors', 'Espais'],
        'Producte': ['Producte', 'Product'],
        'Web': ['Web', 'Digital', 'Interactiu'],
        'Tipografia': ['Tipografia', 'Type'],
        'Fotografia': ['Fotografia', 'Photo'],
        'Il·lustració': ['Il·lustració', 'Illustration']
    }
    
    # Check subject name for itinerari keywords
    subject_lower = subject_name.lower()
    for iti, keywords in itinerari_patterns.items():
        for keyword in keywords:
            if keyword.lower() in subject_lower:
                itinerari = iti
                break
    
    # Look for itinerari headers in nearby cells
    if not itinerari:
        for r in range(max(1, row-5), min(ws.max_row, row+2)):
            for c in range(max(1, col-2), min(ws.max_column, col+3)):
                cell_value = ws.cell(row=r, column=c).value
                if cell_value and isinstance(cell_value, str):
                    for iti, keywords in itinerari_patterns.items():
                        for keyword in keywords:
                            if keyword.lower() in cell_value.lower():
                                itinerari = iti
                                break
    
    return itinerari

def extract_schedule_data(file_path, file_info):
    """Extract all schedule data from an Excel file"""
    
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    extracted_data = []
    
    # Find groups in the file
    groups = []
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                value = str(cell.value).strip()
                # Match group patterns: Gm1, Ga1, M1, A1, etc.
                if re.match(r'^[GMAgma]\d+$', value, re.IGNORECASE):
                    groups.append(value.upper())
    
    # Find semester information
    semester = None
    for row in range(1, min(15, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                value = str(cell.value).strip().lower()
                if '1r semestre' in value or 'primer semestre' in value:
                    semester = 1
                elif '2n semestre' in value or 'segon semestre' in value:
                    semester = 2
    
    # Extract courses with details
    for row in range(5, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            
            # Skip if cell is empty or too short
            if not cell.value or not isinstance(cell.value, str) or len(str(cell.value).strip()) < 3:
                continue
                
            value = str(cell.value).strip()
            
            # Skip patterns that are not subjects
            skip_patterns = [
                'DILLUNS', 'DIMARTS', 'DIMECRES', 'DIJOUS', 'DIVENDRES',
                'SEMESTRE', '09:00', '10:00', '11:00', '11:30', '12:00', '12:30',
                '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00',
                '16:30', '17:00', '17:30', '18:00', '18:30', '19:00', '19:30',
                '20:00', '20:30', 'Tutories', 'CURS', 'NaN', 'lr ', '2n ', 
                '1r ', '3r ', '4t '
            ]
            
            if any(pattern in value for pattern in skip_patterns):
                continue
                
            # Skip if it's just a group code
            if re.match(r'^[GMAgma]\d+$', value, re.IGNORECASE):
                continue
            
            # Check if it's a placeholder (white background)
            is_placeholder = is_white_background(cell)
            
            # Look for professor/classroom info in the next row
            next_cell = ws.cell(row=row+1, column=col) if row < ws.max_row else None
            
            if next_cell and next_cell.value:
                next_value = str(next_cell.value).strip()
                
                # Extract classrooms
                classrooms = re.findall(r'[PGLC][0-9]\.[0-9]+(?:/[0-9]+)?', next_value)
                
                # Also check for special room patterns
                special_rooms = re.findall(r'(?:Platós|Sala\s+\w+|Lab\s+\w+)', next_value)
                classrooms.extend(special_rooms)
                
                if classrooms or len(next_value) > 2:
                    # Remove classroom info to get professor name
                    professor = next_value
                    for classroom in classrooms:
                        professor = professor.replace(classroom, '')
                    
                    # Clean up professor name
                    professor = re.sub(r'\s+', ' ', professor).strip()
                    professor = professor.replace('+', ' ').strip()
                    
                    # Determine group assignment
                    group = None
                    if groups:
                        # Use column position to estimate group
                        if len(groups) == 1:
                            group = groups[0]
                        else:
                            # Estimate based on column position
                            group_index = min((col - 3) // 4, len(groups) - 1)
                            if group_index >= 0 and group_index < len(groups):
                                group = groups[group_index]
                    
                    # Extract itinerari
                    itinerari = extract_itinerari_from_context(ws, row, col, value)
                    
                    # Determine subject type
                    subject_type = file_info.get('tipo', 'Obligatoria')
                    if 'optativ' in value.lower() or 'electiv' in value.lower():
                        subject_type = 'Optativa'
                    elif 'tfg' in value.lower() or 'treball final' in value.lower():
                        subject_type = 'TFG'
                    
                    course_data = {
                        'grado': file_info.get('grado', ''),
                        'grado_code': file_info.get('grado_code', ''),
                        'curso': file_info.get('curso', ''),
                        'semestre': semester,
                        'tipo': subject_type,
                        'itinerari': itinerari,
                        'asignatura': value,
                        'grupo': group or 'Unknown',
                        'aulas': classrooms,
                        'profesor': professor if professor else None,
                        'es_placeholder': is_placeholder,
                        'archivo': os.path.basename(file_path)
                    }
                    
                    # Only add if not a placeholder
                    if not is_placeholder:
                        extracted_data.append(course_data)
    
    return extracted_data

def main():
    """Main function to process all Excel files"""
    all_data = []
    summary = defaultdict(int)
    
    print("Extracting data from Excel files...")
    print("=" * 80)
    
    # Process all Excel files
    for filename in sorted(os.listdir(EXCEL_DIR)):
        if filename.endswith('.xlsx') and not filename.startswith('~'):
            print(f"\nProcessing: {filename}")
            
            file_info = parse_filename(filename)
            file_path = os.path.join(EXCEL_DIR, filename)
            
            try:
                data = extract_schedule_data(file_path, file_info)
                all_data.extend(data)
                
                print(f"  Grado: {file_info.get('grado', 'Unknown')}")
                print(f"  Curso: {file_info.get('curso', 'Unknown')}")
                print(f"  Entries found: {len(data)}")
                
                # Update summary
                for entry in data:
                    summary['total_entries'] += 1
                    summary[f"grado_{entry['grado_code']}"] += 1
                    summary[f"curso_{entry['curso']}"] += 1
                    if entry['itinerari']:
                        summary[f"itinerari_{entry['itinerari']}"] += 1
                    
            except Exception as e:
                print(f"  ERROR: {str(e)}")
    
    # Save to JSON
    output_data = {
        'extraction_date': datetime.now().isoformat(),
        'total_files_processed': len([f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx')]),
        'summary': dict(summary),
        'data': all_data
    }
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    # Print summary
    print("\n" + "=" * 80)
    print("EXTRACTION SUMMARY")
    print("=" * 80)
    print(f"Total entries extracted: {summary['total_entries']}")
    print(f"\nBy degree:")
    for key, value in summary.items():
        if key.startswith('grado_'):
            print(f"  {key.replace('grado_', '')}: {value}")
    print(f"\nBy year:")
    for key, value in summary.items():
        if key.startswith('curso_'):
            print(f"  {key.replace('curso_', '')}: {value}")
    print(f"\nBy itinerari:")
    for key, value in summary.items():
        if key.startswith('itinerari_'):
            print(f"  {key.replace('itinerari_', '')}: {value}")
    
    print(f"\nData saved to: {OUTPUT_FILE}")
    
    # Show sample entries
    print("\nSample entries (first 3):")
    for i, entry in enumerate(all_data[:3]):
        print(f"\n{i+1}. {entry['asignatura']}")
        print(f"   Grado: {entry['grado']} ({entry['grado_code']})")
        print(f"   Curso: {entry['curso']}, Semestre: {entry['semestre']}")
        print(f"   Itinerari: {entry['itinerari'] or 'N/A'}")
        print(f"   Grupo: {entry['grupo']}")
        print(f"   Aulas: {', '.join(entry['aulas']) if entry['aulas'] else 'N/A'}")
        print(f"   Profesor: {entry['profesor'] or 'N/A'}")

if __name__ == "__main__":
    main()