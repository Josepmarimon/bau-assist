#!/usr/bin/env python3
"""
Enhanced extraction script that handles both GDIS and GBA Excel formats.
Version 2: Handles multi-column optatives layout and empty GBA files.
"""

import pandas as pd
import numpy as np
import os
import re
import json
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict

# Directory with Excel files
EXCEL_DIR = '/Users/josepmarimon/Documents/github/bau-assist/csv/excelhorari'
OUTPUT_FILE = '/Users/josepmarimon/Documents/github/bau-assist/csv/extracted_excel_data_v2.json'

def is_white_background(cell):
    """Check if a cell has white background (placeholder)"""
    if cell.fill and cell.fill.patternType:
        if cell.fill.patternType == 'solid':
            if cell.fill.fgColor:
                color = cell.fill.fgColor.rgb
                if color and (color == 'FFFFFFFF' or color == '00FFFFFF' or color == 'FFFFFF'):
                    return True
    return False

def parse_filename(filename):
    """Extract metadata from filename"""
    info = {}
    
    # Determine degree
    if 'GDIS' in filename:
        info['grado'] = 'Disseny'
        info['grado_code'] = 'GDIS'
    elif 'GBA' in filename:
        info['grado'] = 'Belles Arts'
        info['grado_code'] = 'GBA'
    
    # Extract year
    if '_1r_' in filename:
        info['curso'] = '1r'
    elif '_2n_' in filename:
        info['curso'] = '2n'
    elif '_3r_' in filename:
        info['curso'] = '3r'
    elif '_4t_' in filename:
        info['curso'] = '4t'
    elif '3r_4t_Optativitat' in filename:
        info['curso'] = '3r-4t'
        info['tipo'] = 'Optatives'
    
    return info

def extract_subjects_from_multi_option_cell(cell_value):
    """Extract multiple subjects from cells that contain several options separated by lines"""
    if not cell_value or not isinstance(cell_value, str):
        return []
    
    # Split by common separators
    subjects = []
    
    # Split by ----------- separator
    parts = cell_value.split('-----------')
    
    for part in parts:
        part = part.strip()
        if not part:
            continue
            
        # Extract subject name and professor/classroom
        lines = part.split('\n')
        if lines:
            subject_name = lines[0].strip()
            # Skip if it's just a classroom or too short
            if (len(subject_name) > 5 and 
                not re.match(r'^[PGL]\d+\.\d+', subject_name) and
                subject_name not in ['tutories', 'Tutories']):
                
                # Extract professor and classroom from remaining lines
                professor = None
                classrooms = []
                
                for line in lines[1:]:
                    line = line.strip()
                    # Extract classrooms
                    found_classrooms = re.findall(r'[PGL]\d+\.\d+(?:/\d+)?', line)
                    classrooms.extend(found_classrooms)
                    
                    # The rest might be professor name
                    prof_text = line
                    for classroom in found_classrooms:
                        prof_text = prof_text.replace(classroom, '')
                    prof_text = prof_text.strip()
                    
                    if prof_text and len(prof_text) > 3:
                        professor = prof_text
                
                subjects.append({
                    'name': subject_name,
                    'professor': professor,
                    'classrooms': classrooms
                })
    
    return subjects

def extract_gba_optatives_data(file_path, file_info):
    """Special extraction for GBA optatives file with multi-column layout"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    extracted_data = []
    
    # Process the specific structure of optatives
    # Columns 3, 5, 6, and 9 contain subject information
    subject_columns = [3, 5, 6, 9]
    
    for row in range(4, 30):  # Typical range where subjects appear
        for col in subject_columns:
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                value = str(cell.value).strip()
                
                # Skip common non-subject values
                if value in ['tutories', 'tutories P0.10', '1r semestre', '2n semestre'] or len(value) < 5:
                    continue
                
                # Handle multi-option cells
                if '-----------' in value or '\n' in value:
                    subjects = extract_subjects_from_multi_option_cell(value)
                    for subj in subjects:
                        course_data = {
                            'grado': file_info.get('grado', ''),
                            'grado_code': file_info.get('grado_code', ''),
                            'curso': file_info.get('curso', ''),
                            'semestre': 1 if row < 16 else 2,  # Row 16 is typically where 2nd semester starts
                            'tipo': 'Optativa',
                            'itinerari': None,
                            'asignatura': subj['name'],
                            'grupo': 'Opt',
                            'aulas': subj['classrooms'],
                            'profesor': subj['professor'],
                            'es_placeholder': False,
                            'archivo': os.path.basename(file_path)
                        }
                        extracted_data.append(course_data)
                else:
                    # Single subject entry
                    # Check if next row has professor info
                    next_cell = ws.cell(row=row+1, column=col) if row < ws.max_row else None
                    professor = None
                    classrooms = []
                    
                    if next_cell and next_cell.value:
                        next_value = str(next_cell.value).strip()
                        classrooms = re.findall(r'[PGL]\d+\.\d+(?:/\d+)?', next_value)
                        
                        # Extract professor
                        professor = next_value
                        for classroom in classrooms:
                            professor = professor.replace(classroom, '')
                        professor = professor.strip()
                        
                        if len(professor) < 3:
                            professor = None
                    
                    # Only add if it looks like a subject
                    if not re.match(r'^[PGL]\d+\.\d+', value) and value not in ['d\'Artista', 'Pedagogies', 'Experimentals', 'Guarin']:
                        course_data = {
                            'grado': file_info.get('grado', ''),
                            'grado_code': file_info.get('grado_code', ''),
                            'curso': file_info.get('curso', ''),
                            'semestre': 1 if row < 16 else 2,
                            'tipo': 'Optativa',
                            'itinerari': None,
                            'asignatura': value,
                            'grupo': 'Opt',
                            'aulas': classrooms,
                            'profesor': professor,
                            'es_placeholder': False,
                            'archivo': os.path.basename(file_path)
                        }
                        extracted_data.append(course_data)
    
    return extracted_data

def extract_schedule_data(file_path, file_info):
    """Extract schedule data - use special handling for GBA optatives"""
    
    # Special handling for GBA optatives file
    if 'Optativitat' in os.path.basename(file_path):
        return extract_gba_optatives_data(file_path, file_info)
    
    # Check if it's an empty GBA file
    if file_info.get('grado_code') == 'GBA':
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Count non-empty cells
        non_empty = 0
        for row in range(1, min(30, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value and str(cell.value).strip():
                    non_empty += 1
        
        # If file has very few non-empty cells, it's probably empty
        if non_empty < 20:
            print(f"    File appears to be empty (only {non_empty} non-empty cells)")
            return []
    
    # Original extraction logic for GDIS files
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    extracted_data = []
    
    # Find groups in the file
    groups = []
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                value = str(cell.value).strip()
                if re.match(r'^[GMAgma]\d+$', value, re.IGNORECASE):
                    groups.append(value.upper())
    
    # Find semester information
    semester = None
    for row in range(1, min(15, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                value = str(cell.value).strip().lower()
                if '1r semestre' in value or 'primer semestre' in value:
                    semester = 1
                elif '2n semestre' in value or 'segon semestre' in value:
                    semester = 2
    
    # Extract courses with details
    for row in range(5, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            
            if not cell.value or not isinstance(cell.value, str) or len(str(cell.value).strip()) < 3:
                continue
                
            value = str(cell.value).strip()
            
            # Skip patterns
            skip_patterns = [
                'DILLUNS', 'DIMARTS', 'DIMECRES', 'DIJOUS', 'DIVENDRES',
                'SEMESTRE', '09:00', '10:00', '11:00', '11:30', '12:00', '12:30',
                '13:00', '13:30', '14:00', '14:30', '15:00', '15:30', '16:00',
                '16:30', '17:00', '17:30', '18:00', '18:30', '19:00', '19:30',
                '20:00', '20:30', 'Tutories', 'CURS', 'NaN', 'lr ', '2n ', 
                '1r ', '3r ', '4t '
            ]
            
            if any(pattern in value for pattern in skip_patterns):
                continue
                
            if re.match(r'^[GMAgma]\d+$', value, re.IGNORECASE):
                continue
            
            is_placeholder = is_white_background(cell)
            
            next_cell = ws.cell(row=row+1, column=col) if row < ws.max_row else None
            
            if next_cell and next_cell.value:
                next_value = str(next_cell.value).strip()
                
                classrooms = re.findall(r'[PGLC][0-9]\.[0-9]+(?:/[0-9]+)?', next_value)
                special_rooms = re.findall(r'(?:Platós|Sala\s+\w+|Lab\s+\w+)', next_value)
                classrooms.extend(special_rooms)
                
                if classrooms or len(next_value) > 2:
                    professor = next_value
                    for classroom in classrooms:
                        professor = professor.replace(classroom, '')
                    
                    professor = re.sub(r'\s+', ' ', professor).strip()
                    professor = professor.replace('+', ' ').strip()
                    
                    group = None
                    if groups:
                        if len(groups) == 1:
                            group = groups[0]
                        else:
                            group_index = min((col - 3) // 4, len(groups) - 1)
                            if group_index >= 0 and group_index < len(groups):
                                group = groups[group_index]
                    
                    # Extract itinerari
                    itinerari = None
                    subject_lower = value.lower()
                    itinerari_patterns = {
                        'Videojocs': ['videojocs', 'game', 'jocs'],
                        'Animació': ['animació', 'animation'],
                        'Audiovisual': ['audiovisual', 'vídeo', 'cinema', 'llenguatges audiovisuals'],
                        'Gràfic': ['gràfic', 'comunicació visual', 'disseny gràfic', 'expressió gràfica'],
                        'Moda': ['moda', 'fashion'],
                        'Interiors': ['interiors', 'espais'],
                        'Producte': ['producte', 'product'],
                        'Web': ['web', 'digital', 'interactiu', 'creació i autoria digital'],
                        'Tipografia': ['tipografia', 'type'],
                        'Fotografia': ['fotografia', 'photo'],
                        'Il·lustració': ['il·lustració', 'illustration']
                    }
                    
                    for iti, keywords in itinerari_patterns.items():
                        for keyword in keywords:
                            if keyword in subject_lower:
                                itinerari = iti
                                break
                    
                    subject_type = file_info.get('tipo', 'Obligatoria')
                    if 'optativ' in value.lower() or 'electiv' in value.lower():
                        subject_type = 'Optativa'
                    elif 'tfg' in value.lower() or 'treball final' in value.lower():
                        subject_type = 'TFG'
                    
                    course_data = {
                        'grado': file_info.get('grado', ''),
                        'grado_code': file_info.get('grado_code', ''),
                        'curso': file_info.get('curso', ''),
                        'semestre': semester,
                        'tipo': subject_type,
                        'itinerari': itinerari,
                        'asignatura': value,
                        'grupo': group or 'Unknown',
                        'aulas': classrooms,
                        'profesor': professor if professor else None,
                        'es_placeholder': is_placeholder,
                        'archivo': os.path.basename(file_path)
                    }
                    
                    if not is_placeholder:
                        extracted_data.append(course_data)
    
    return extracted_data

def main():
    """Main function to process all Excel files"""
    all_data = []
    summary = defaultdict(int)
    
    print("Extracting data from Excel files (v2 - Enhanced for GBA)...")
    print("=" * 80)
    
    # Process all Excel files
    for filename in sorted(os.listdir(EXCEL_DIR)):
        if filename.endswith('.xlsx') and not filename.startswith('~'):
            print(f"\nProcessing: {filename}")
            
            file_info = parse_filename(filename)
            file_path = os.path.join(EXCEL_DIR, filename)
            
            try:
                data = extract_schedule_data(file_path, file_info)
                all_data.extend(data)
                
                print(f"  Grado: {file_info.get('grado', 'Unknown')}")
                print(f"  Curso: {file_info.get('curso', 'Unknown')}")
                print(f"  Entries found: {len(data)}")
                
                # Update summary
                for entry in data:
                    summary['total_entries'] += 1
                    summary[f"grado_{entry['grado_code']}"] += 1
                    summary[f"curso_{entry['curso']}"] += 1
                    if entry['itinerari']:
                        summary[f"itinerari_{entry['itinerari']}"] += 1
                    
            except Exception as e:
                print(f"  ERROR: {str(e)}")
    
    # Save to JSON
    output_data = {
        'extraction_date': datetime.now().isoformat(),
        'version': 2,
        'total_files_processed': len([f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx')]),
        'summary': dict(summary),
        'data': all_data
    }
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    # Print summary
    print("\n" + "=" * 80)
    print("EXTRACTION SUMMARY (V2)")
    print("=" * 80)
    print(f"Total entries extracted: {summary['total_entries']}")
    print(f"\nBy degree:")
    for key, value in summary.items():
        if key.startswith('grado_'):
            print(f"  {key.replace('grado_', '')}: {value}")
    print(f"\nBy year:")
    for key, value in summary.items():
        if key.startswith('curso_'):
            print(f"  {key.replace('curso_', '')}: {value}")
    print(f"\nBy itinerari:")
    for key, value in summary.items():
        if key.startswith('itinerari_'):
            print(f"  {key.replace('itinerari_', '')}: {value}")
    
    print(f"\nData saved to: {OUTPUT_FILE}")
    
    # Show GBA entries
    gba_entries = [e for e in all_data if e['grado_code'] == 'GBA']
    print(f"\nBelles Arts entries found: {len(gba_entries)}")
    if gba_entries:
        print("\nSample GBA entries:")
        for i, entry in enumerate(gba_entries[:5]):
            print(f"\n{i+1}. {entry['asignatura']}")
            print(f"   Curso: {entry['curso']}, Tipo: {entry['tipo']}")
            print(f"   Aulas: {', '.join(entry['aulas']) if entry['aulas'] else 'N/A'}")
            print(f"   Profesor: {entry['profesor'] or 'N/A'}")

if __name__ == "__main__":
    main()